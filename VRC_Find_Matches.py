from Tkinter import Tk, Label, StringVar, Entry, N, W
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Fill, colors
from openpyxl.cell import Cell
from dateutil import tz, parser
import requests
import tkMessageBox
from sets import Set

class guiInterface:
    def __init__(self,master):
        # set title
        self.master = master
        master.title('Import Matches')

        # set instruction labels
        self.instructionLabel = Label(master,text='Fill in parameters and press \'enter\' to search for events')
        self.teamLabel = Label(master, text = 'Team Number:')
        self.divisionLabel = Label(master, text = 'Division: (For larger events)')

        # set result labels
        self.resultTexts = []
        self.resultLabels = []
        for x in xrange(6):
            aResultText = StringVar()
            aResultLabel = Label(master,textvariable=aResultText)
            aResultLabel.grid(row=x,column=2)
            self.resultTexts.append(aResultText)
            self.resultLabels.append(aResultLabel)

        # set data entry for date and team number
        self.divisionEntry = Entry(master)
        self.teamEntry = Entry(master)
        master.bind('<Return>',self.getEvents)

        # LAYOUT
        self.instructionLabel.grid(row=0,column=0,rowspan=2, sticky=N)
        self.teamLabel.grid(row=2,column=0, sticky=W)
        self.teamEntry.grid(row=3,column=0, sticky=W)
        self.divisionLabel.grid(row=4,column=0, sticky=W)
        self.divisionEntry.grid(row=5,column=0, sticky=W)
        master.grid_columnconfigure(1,minsize=50)
        master.grid_columnconfigure(2,minsize=500)
        master.grid_rowconfigure(7,minsize=50)

    # Update the event list that fulfills the given parameters
    def getEvents(self,event):
            # get parameters
            self.team = self.teamEntry.get()

            # make call to vexDB api
            findEventUrl = 'https://api.vexdb.io/v1/get_events'
            findEventParams = {'team':self.team}
            self.eventResponse = requests.get(findEventUrl,findEventParams).json()

            # limit number of events to 6 at most
            loopEnd = 0
            if(self.eventResponse['size']<7):
                loopEnd = self.eventResponse['size']
            else:
                loopEnd = 6

            # clear event list
            for _ in range(6):
                self.resultTexts[_].set('')

            # update event list
            for eventIndex in range(loopEnd):
                x = int(eventIndex)
                self.resultTexts[eventIndex].set(self.eventResponse['result'][eventIndex]['name'])
                self.resultLabels[eventIndex].bind('<Enter>',lambda event: event.widget.configure(font='Lucida 13 bold'))
                self.resultLabels[eventIndex].bind('<Leave>',lambda event: event.widget.configure(font='Lucida 13'))
                self.resultLabels[eventIndex].bind('<Button-1>',self.findMatches)

    # find matches
    def findMatches(self,event):
        # Display status
        self.status_text = StringVar()
        self.status_text.set('Importing matches...')
        self.status_label = Label(self.master, textvariable = self.status_text)
        self.status_label.grid(row=7,column=0,columnspan=3)
        Tk.update(self.master)

        # Define set to record teams in the event
        teams = Set()
        opponentsOne = []
        opponentsTwo = []
        allies = []
        matchIndexes = []

        # find parameters (sku and division)
        eventIndex = self.resultLabels.index(event.widget)
        sku = self.eventResponse['result'][eventIndex]['sku']
        division = self.divisionEntry.get()

        # call vexDB api for matches
        url = 'https://api.vexdb.io/v1/get_matches'
        params = {'sku':sku,'round':2,'division':division}
        matchResults = requests.get(url,params).json()

        # write matches to excel file
        # define variables
        filename = 'VRC_Scouting.xlsm'
        wb = load_workbook(filename=filename, read_only=False, keep_vba=True)
        wsMatches = wb['MatchData']
        wsTeams = wb['TeamData']

        # check to see if there is unsaved data
        if(wsMatches['A3'].value != None or wsTeams['A3'].value != None):
            proceed = tkMessageBox.askokcancel('Import Matches','Warning: The current contexts of the sheet will be deleted')
            if(not proceed):
                return

        # reset cells and formatting
        for row in range(3,wsMatches.max_row+1): #for match schedule
            for column in range(1,wsMatches.max_column+1):
                cell = wsMatches.cell(row=row,column=column)
                cell.value = None
                cell.style = 'Normal'
                cell.font = Font(name = 'Times New Roman')
                # cell.fill = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')

        for row in range(3,wsTeams.max_row+1): #for team list
            for column in range(1,wsTeams.max_column+1):
                cell = wsTeams.cell(row=row,column=column)
                cell.value = None

        # write new match data and record teams
        for matchIndex in range(matchResults['size']):
            wsMatches.cell(row=matchIndex+3,column=1).value = 'Q' + str(matchResults['result'][matchIndex]['matchnum'])
            wsMatches.cell(row=matchIndex+3,column=2).value = matchResults['result'][matchIndex]['field']
            # format match time
            matchTime = parser.parse(matchResults['result'][matchIndex]['scheduled'])
            matchTime.replace(tzinfo=tz.gettz('UTC'))
            wsMatches.cell(row=matchIndex+3,column=3).value = matchTime.astimezone(tz.gettz('EST')).strftime('%I:%M %p')
            wsMatches.cell(row=matchIndex+3,column=4).value = matchResults['result'][matchIndex]['red1']
            wsMatches.cell(row=matchIndex+3,column=5).value = matchResults['result'][matchIndex]['red2']
            wsMatches.cell(row=matchIndex+3,column=6).value = matchResults['result'][matchIndex]['blue1']
            wsMatches.cell(row=matchIndex+3,column=7).value = matchResults['result'][matchIndex]['blue2']

            # check if match contains team -> if so, highlight that match appropriately and record opponents/allies
            match = matchResults['result'][matchIndex]
            if self.team in [match['red1'], match['red2'], match['blue1'], match['blue2']]:
                if self.team in [match['red1'], match['red2']]: #if red, highlight each cell in the row red
                    for col in xrange(1,8):
                        wsMatches.cell(row=matchIndex+3,column=col).fill = PatternFill(fill_type='solid', start_color='ffffafaf', end_color='ffffafaf')
                else: #if blue, highlight each cell in the row blue
                    for col in xrange(1,8):
                        wsMatches.cell(row=matchIndex+3,column=col).fill = PatternFill(fill_type='solid', start_color='ffa5a5ff', end_color='ffa5a5ff')

                # update allies and opponents
                if self.team == match['red1']:
                    allies.append(matchResults['result'][matchIndex]['red2'])
                    opponentsOne.append(matchResults['result'][matchIndex]['blue1'])
                    opponentsTwo.append(matchResults['result'][matchIndex]['blue2'])
                elif self.team == match['red2']:
                    allies.append(matchResults['result'][matchIndex]['red1'])
                    opponentsOne.append(matchResults['result'][matchIndex]['blue1'])
                    opponentsTwo.append(matchResults['result'][matchIndex]['blue2'])
                elif self.team == match['blue1']:
                    allies.append(matchResults['result'][matchIndex]['blue2'])
                    opponentsOne.append(matchResults['result'][matchIndex]['red1'])
                    opponentsTwo.append(matchResults['result'][matchIndex]['red2'])
                else:
                    allies.append(matchResults['result'][matchIndex]['blue1'])
                    opponentsOne.append(matchResults['result'][matchIndex]['red1'])
                    opponentsTwo.append(matchResults['result'][matchIndex]['red2'])

                # record index of match
                matchIndexes.append(matchIndex)

            # add teams to teams list
            teams.add(matchResults['result'][matchIndex]['red1'])
            teams.add(matchResults['result'][matchIndex]['red2'])
            teams.add(matchResults['result'][matchIndex]['blue1'])
            teams.add(matchResults['result'][matchIndex]['blue2'])

        # highlight allies and enemies
        for row in xrange(matchResults['size']):
            match = matchResults['result'][row]
            # Find the games after the current row
            firstGame = 0;
            while(firstGame<len(matchIndexes) and matchIndexes[firstGame]<=(row+3)): #+3 assumes that rows start at 3
                firstGame += 1

            for col in xrange(4): #iterate over teams in the row
                if col<2:
                    teamStr = 'red' + str(col+1)
                else:
                    teamStr = 'blue' + str(col-1)

                # find allies/opponents in games after current row and highlight appropriately
                for x in xrange(firstGame,len(matchIndexes)):
                    if match[teamStr] == allies[x]:
                        wsMatches.cell(row=row+3,column=col+4).fill = PatternFill(fill_type='solid', start_color='ffd7ff00', end_color='ffd7ff00')
                    elif match[teamStr] == opponentsOne[x] or match[teamStr] == opponentsTwo[x]:
                        wsMatches.cell(row=row+3,column=col+4).fill = PatternFill(fill_type='solid', start_color='ffffd300', end_color='ffffd300')



        # write teams
        rowIndex = 3
        for team in teams:
            wsTeams.cell(row=rowIndex,column=1).value = team
            rowIndex += 1

        # mark tournament and team name on excel sheet
        wsMatches['L1'] = self.teamEntry.get()
        wsMatches['L2'] = self.eventResponse['result'][eventIndex]['name']

        # save changes
        wb.save(filename)

        # signal end
        self.status_text.set('Finished importing matches')
        Tk.update(self.master)


# Used to center window, code taken from https://stackoverflow.com/questions/3352918/how-to-center-a-window-on-the-screen-in-tkinter
def center(toplevel):
    toplevel.update_idletasks()
    w = toplevel.winfo_screenwidth()
    h = toplevel.winfo_screenheight()
    size = tuple(int(_) for _ in toplevel.geometry().split('+')[0].split('x'))
    x = w/2 - size[0]/2
    y = h/2 - size[1]/2
    toplevel.geometry("%dx%d+%d+%d" % (size + (x, y)))

# create window
root = Tk()
app = guiInterface(root)
center(root)
root.mainloop()

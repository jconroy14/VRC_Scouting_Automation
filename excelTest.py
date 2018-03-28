import os
from openpyxl import *

dir_path = os.path.dirname(os.path.realpath(__file__))
cwd = os.getcwd()

wb = load_workbook(filename='VRC_Scouting.xlsm', read_only=False, keep_vba=True)
print str(wb.sheetnames)

ws = wb['MatchData']
print str(ws.title)

ws['A2'].value = 'hi'
print(ws['A2'].value)

matchDataToAdd = [
[1,2,3],
[1,'hi',3,4],
[1,2,3,4,5]
]

print ws.max_row
print ws.max_column

for row in range(3,ws.max_row+1):
    for column in range(1,ws.max_column+1):
        cell = ws.cell(row=row,column=column)
        cell.value = None
        cell.style = 'Normal'

for row in matchDataToAdd:
    ws.append(row)

wb.save('VRC_Scouting.xlsm')
